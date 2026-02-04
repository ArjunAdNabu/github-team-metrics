"""Excel export functionality with formatting."""

import logging
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger('github_metrics')


class ExcelExporter:
    """Export data to Excel with formatting."""

    def __init__(self, output_path: str):
        """
        Initialize Excel exporter.

        Args:
            output_path: Path where Excel file will be saved
        """
        self.output_path = output_path
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def create_summary_sheet(self, data: List[Dict]):
        """
        Create summary worksheet with overview statistics.

        Args:
            data: List of combined user metrics
        """
        ws = self.workbook.create_sheet('Summary', 0)

        # Title
        ws['A1'] = 'Team Performance Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')

        # Team-wide metrics
        total_commits = sum(d.get('total_commits', 0) for d in data)
        total_lines_added = sum(d.get('lines_added', 0) for d in data)
        total_lines_deleted = sum(d.get('lines_deleted', 0) for d in data)
        total_prs = sum(d.get('prs_created', 0) for d in data)
        total_issues_closed = sum(d.get('issues_closed', 0) for d in data)
        total_complexity_score = sum(d.get('total_complexity_score', 0) for d in data)
        total_reviews = sum(d.get('reviews_given', 0) for d in data)
        total_tickets = sum(d.get('total_tickets', 0) for d in data)
        total_tickets_closed = sum(d.get('tickets_closed', 0) for d in data)
        total_sla_failures = sum(d.get('sla_failures', 0) for d in data)

        row = 3
        metrics = [
            ('Total Team Members', len(data)),
            ('Total Commits', total_commits),
            ('Total Lines Added', total_lines_added),
            ('Total Lines Deleted', total_lines_deleted),
            ('Total Pull Requests', total_prs),
            ('Total Issues Closed', total_issues_closed),
            ('Total Complexity Score', round(total_complexity_score, 1)),
            ('Total Code Reviews', total_reviews),
            ('Total Tickets', total_tickets),
            ('Tickets Closed', total_tickets_closed),
            ('SLA Failures (>48 business hrs)', total_sla_failures),
            ('Team SLA Success Rate %', round(((total_tickets_closed - total_sla_failures) / total_tickets_closed * 100), 1) if total_tickets_closed > 0 else 100.0),
            ('Average Commits per Person', round(total_commits / len(data), 1) if data else 0),
            ('Average PRs per Person', round(total_prs / len(data), 1) if data else 0),
            ('Average Complexity Score per Person', round(total_complexity_score / len(data), 1) if data else 0),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Top contributors
        row += 2
        ws[f'A{row}'] = 'Top Contributors'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Sort by activity score
        sorted_data = sorted(data, key=lambda x: x.get('activity_score', 0), reverse=True)

        ws[f'A{row}'] = 'Name'
        ws[f'B{row}'] = 'Commits'
        ws[f'C{row}'] = 'PRs'
        ws[f'D{row}'] = 'Reviews'
        ws[f'E{row}'] = 'Tickets'
        ws[f'F{row}'] = 'Complexity'
        ws[f'G{row}'] = 'Activity Score'

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color='D3D3D3', fill_type='solid')

        row += 1

        for user in sorted_data[:10]:  # Top 10
            ws[f'A{row}'] = user.get('display_name', '')
            ws[f'B{row}'] = user.get('total_commits', 0)
            ws[f'C{row}'] = user.get('prs_created', 0)
            ws[f'D{row}'] = user.get('reviews_given', 0)
            ws[f'E{row}'] = user.get('total_tickets', 0)
            ws[f'F{row}'] = user.get('total_complexity_score', 0)
            ws[f'G{row}'] = user.get('activity_score', 0)
            row += 1

        # Auto-size columns
        self._auto_size_columns(ws)

    def create_team_metrics_sheet(self, data: List[Dict]):
        """
        Create detailed team metrics worksheet.

        Args:
            data: List of combined user metrics
        """
        ws = self.workbook.create_sheet('Team Metrics')

        # Define columns
        columns = [
            ('GitHub Username', 'github_username'),
            ('Display Name', 'display_name'),
            ('Email', 'email'),
            ('Commits', 'total_commits'),
            ('Commit Freq', 'commit_frequency'),
            ('Lines Added', 'lines_added'),
            ('Lines Deleted', 'lines_deleted'),
            ('Lines Changed', 'lines_changed'),
            ('PRs Created', 'prs_created'),
            ('PRs Merged', 'prs_merged'),
            ('PR Merge Rate %', 'pr_merge_rate'),
            ('Avg PR Size', 'avg_pr_size'),
            ('Issues Closed', 'issues_closed'),
            ('Complexity Score', 'total_complexity_score'),
            ('Reviews Given', 'reviews_given'),
            ('Reviews Received', 'reviews_received'),
            ('Review Participation', 'review_participation'),
            ('Avg Review Time (hrs)', 'avg_review_time_hours'),
            ('Tickets', 'total_tickets'),
            ('Tickets Open', 'tickets_open'),
            ('Tickets Closed', 'tickets_closed'),
            ('High Priority', 'tickets_high_priority'),
            ('Medium Priority', 'tickets_medium_priority'),
            ('Low Priority', 'tickets_low_priority'),
            ('Avg Resolution (hrs)', 'avg_resolution_time_hours'),
            ('Avg Business Resolution (hrs)', 'avg_business_resolution_hours'),
            ('Avg First Response (hrs)', 'avg_first_response_time_hours'),
            ('SLA Failures (>48h)', 'sla_failures'),
            ('SLA Success Rate %', 'sla_success_rate'),
            ('Tickets w/ GitHub Issue', 'tickets_with_github_issue'),
            ('Commits/Ticket', 'commits_per_ticket'),
            ('Activity Score', 'activity_score'),
            ('Active Repos', 'active_repos'),
            ('Data Sources', 'data_sources'),
        ]

        # Write headers
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row_idx, user in enumerate(data, 2):
            for col_idx, (_, key) in enumerate(columns, 1):
                value = user.get(key, '')
                ws.cell(row_idx, col_idx, value)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add filters
        ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'

        # Auto-size columns
        self._auto_size_columns(ws)

    def create_repository_breakdown_sheet(self, repos: List[Dict], all_commits: List[Dict]):
        """
        Create repository breakdown worksheet.

        Args:
            repos: List of repository info
            all_commits: All commits data
        """
        ws = self.workbook.create_sheet('Repository Breakdown')

        # Headers
        headers = ['Repository', 'Commits', 'Contributors', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')

        # Calculate repo metrics
        repo_metrics = {}
        for commit in all_commits:
            repo_name = commit.get('repo', '')
            if repo_name not in repo_metrics:
                repo_metrics[repo_name] = {
                    'commits': 0,
                    'contributors': set()
                }

            repo_metrics[repo_name]['commits'] += 1

            author = commit.get('author', {}).get('user')
            if author and author.get('login'):
                repo_metrics[repo_name]['contributors'].add(author['login'])

        # Write data
        row = 2
        for repo in repos:
            repo_name = repo.get('nameWithOwner', '')
            metrics = repo_metrics.get(repo_name, {'commits': 0, 'contributors': set()})

            ws.cell(row, 1, repo_name)
            ws.cell(row, 2, metrics['commits'])
            ws.cell(row, 3, len(metrics['contributors']))
            ws.cell(row, 4, 'Archived' if repo.get('isArchived') else 'Active')

            row += 1

        # Auto-size columns
        self._auto_size_columns(ws)

    def create_ticket_details_sheet(self, tickets: List[Dict]):
        """
        Create ticket details worksheet.

        Args:
            tickets: List of individual ticket dictionaries
        """
        if not tickets:
            logger.warning("No tickets to export")
            return

        ws = self.workbook.create_sheet('Ticket Details')

        # Headers (13 columns from the source data)
        headers = [
            'Title',
            'Priority',
            'Type',
            'Assigned',
            'Reported by',
            'Reported Time',
            'First Response Time',
            'Closed Time',
            'Duration',
            'Bucket',
            'GitHub Issue',
            'Notes',
            'Root Cause Status'
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')

        # Write data
        for row_idx, ticket in enumerate(tickets, 2):
            ws.cell(row_idx, 1, ticket.get('title', ''))
            ws.cell(row_idx, 2, ticket.get('priority', ''))
            ws.cell(row_idx, 3, ticket.get('type', ''))
            ws.cell(row_idx, 4, ticket.get('assigned', ''))
            ws.cell(row_idx, 5, ticket.get('reported_by', ''))

            # Dates
            reported_time = ticket.get('reported_time')
            ws.cell(row_idx, 6, reported_time.strftime('%Y-%m-%d %H:%M') if reported_time else '')

            first_response = ticket.get('first_response_time')
            ws.cell(row_idx, 7, first_response.strftime('%Y-%m-%d %H:%M') if first_response else '')

            closed_time = ticket.get('closed_time')
            ws.cell(row_idx, 8, closed_time.strftime('%Y-%m-%d %H:%M') if closed_time else '')

            ws.cell(row_idx, 9, ticket.get('duration', ''))
            ws.cell(row_idx, 10, ticket.get('bucket', ''))
            ws.cell(row_idx, 11, ticket.get('github_issue', ''))
            ws.cell(row_idx, 12, ticket.get('notes', ''))
            ws.cell(row_idx, 13, ticket.get('root_cause_status', ''))

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add filters
        ws.auto_filter.ref = f'A1:M1'

        # Auto-size columns
        self._auto_size_columns(ws)

    def _auto_size_columns(self, worksheet):
        """
        Auto-size all columns based on content.

        Args:
            worksheet: Worksheet to adjust
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def apply_formatting(self):
        """Apply consistent formatting to all sheets."""
        # This is called after all sheets are created
        for sheet in self.workbook:
            # Add borders to all cells with data
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.border = Border(
                            left=Side(style='thin', color='D3D3D3'),
                            right=Side(style='thin', color='D3D3D3'),
                            top=Side(style='thin', color='D3D3D3'),
                            bottom=Side(style='thin', color='D3D3D3')
                        )

    def save(self):
        """Save workbook to file."""
        try:
            self.workbook.save(self.output_path)
            logger.info(f"Excel file saved: {self.output_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise
